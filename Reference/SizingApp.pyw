#TODO: Have user choose default folders -  LISTS folder and Save folder. 
#TODO: Door layout preview - draw image of door on table. 
#TODO: Break filenames out by job/room
#TODO: Ability to use different tool numbers. 

#import xlrd
import sys, os, re
import openpyxl
import ConfigParser as con

from PyQt4 import QtCore, QtGui
from door_sizing_ui import Ui_Dialog as doorSizing_Ui
from config_ui import Ui_Dialog as config_Ui

#Global data. Modify these if needed

toolDia = 2.5312                    # Diameter of cutting bit. 
toolRadius = toolDia/2              # This is the magic number.    #was 1.2646
doorCount = 0

LISTS_Folder = 'O:\\OFFICE\\JOBS 2014'
Output_Folder = ''
Oversize_Amount = 0
Default_Tool = 2

cfg=con.ConfigParser()

door = []
 
#For debugging purposes...
sys.stdout = open("c:\\sizing_debug.log", "w")

if os.path.isfile("c:\\doorsizing.ini"):
	cfg.read("C:\\doorsizing.ini")
	Oversize_Amount=cfg.get("Options","oversize_amount")
	LISTS_Folder=cfg.get("Options","lists_folder")
	Output_Folder=cfg.get("Options","output_folder")
	Default_Tool=cfg.get("Options","default_tool")
else:
	cfg.add_section("Options")
	cfg.set("Options","oversize_amount",Oversize_Amount)              
	cfg.set("Options","lists_folder",LISTS_Folder)                    
	cfg.set("Options","output_folder",Output_Folder) 
	cfg.set("Options","default_tool",Default_Tool)               
	dummy=open("C:\\doorsizing.ini","w")
	cfg.write(dummy)
	dummy.close()

class doorClass:
	def __init__(self, pNum=0, width=0, height=0, toolNum=Default_Tool, lineNum='', jobName='', page=''):
		self.width = width
		self.height = height
		self.jobName = jobName
		#self.roomName = roomName
		self.lineNum = lineNum
		self.pNum = pNum
		self.page = page
		self.toolNum = toolNum

	def get_pNum(self):
		try:
			return int(self.pNum)
		except:
			try: 
				return float(self.pNum)
			except:
				return str(self.pNum)

	def gen_file(self, path=''):
		global toolRadius,Oversize_Amount,toolDia
		dName = path + self.jobName + "/" + self.page

		if not os.path.exists(dName):
			try:
				os.makedirs(dName)

			except Exception as e:
				print "DEBUG: An error occured creating the path... "
				print repr(e)

		fileName = dName + '/' + str(self.page) + ' #' + str(self.get_pNum()) + ' - Line '+ str(self.lineNum) + '.out'
		
		if fileName:
			fileHeader = [  ":1", 
								"("+fileName+")", 
								"N1 G00 G17 G20 G40 G49 G64 G80 G90 Z0 M5", 
								"N2 G52 X0 Y0 Z0", 
								"N3 G08 P1", 
								"N4 (ROUTER-BIT GENERIC SIZE)", 
								"N5 G49 Z0 M05", 
								"N6 G90 T200" + str(self.toolNum) +" M06", 
								"N7 T102", 
								"N8 M03 S18000"]
										
			fileFooter = [  "N22 G00 Z.25", 
								"N23 G40 G00 Y-5.1881", 
								"N24 G49 Z0 M5", 
								"N25 G28 G91 X0 Y0", 
								"N26 G52 X0 Y0 Z0", 
								"N27 G08 P0", 
								"N28 M30", 
								"%"]

			self.outFile = open(fileName, "w")
			
			if Oversize_Amount > 0: 
				oa = (float(Oversize_Amount)/2)
			else:
				oa = 0.0

			#Insert Header
			for x in range(0, len(fileHeader)):         
					print >> self.outFile,  fileHeader[x]
			#Begin/Lead In Code
			print >> self.outFile, "N9 G00 G17 G55 X" + str((self.height/2)-toolDia)+" Y-5.1881"
			print >> self.outFile, "N10 G00 G43 H2 Z.25"
			print >> self.outFile, "N11 Z.1"
			print >> self.outFile, "N12 G42 D12 G01 Y-3.7959 F75."
			print >> self.outFile, "N13 Z0."
			print >> self.outFile, "N14 G02 X"+ str(self.height/2)+" Y" + str(0-toolRadius-oa) + " I2.5313 F150."
			#Cut out code
			print >> self.outFile, "N15 G01 X" + str(self.height+toolRadius+oa)
			print >> self.outFile, "N16 Y"+ str(self.width+toolRadius+oa)
			print >> self.outFile, "N17 X"+ str(0-toolRadius-oa)
			print >> self.outFile, "N18 Y"+ str(0-toolRadius-oa)
			print >> self.outFile, "N19 X"+ str(self.height/2)
			print >> self.outFile, "N20 X"+ str((self.height/2)+toolDia) +" F150."
			#Lead out
			print >> self.outFile, "N21 G02 X"+ str(str((self.height/2)+toolDia+toolDia)) +" Y-3.7959 J-2.5313"
			#Insert Footer
			for x in range(0, len(fileFooter)):
					print >> self.outFile,  fileFooter[x]
			self.outFile.close()

class ConfigDialog(QtGui.QDialog):
	def __init__(self, parent=None):
		QtGui.QWidget.__init__(self, parent)
		global Oversize_Amount,LISTS_Folder,Output_Folder,Default_Tool
		self.ui = config_Ui()
		self.ui.setupUi(self)
		self.setFixedSize(479,205)
		self.setWindowFlags(QtCore.Qt.CustomizeWindowHint)
		self.ui.oversizeBox.setValue(float(Oversize_Amount))
		self.ui.listsEdit.setText(LISTS_Folder)
		self.ui.outputEdit.setText(Output_Folder)
		self.ui.toolNumBox.setValue(int(Default_Tool))
		self.connect(self.ui.cancelButton, QtCore.SIGNAL("clicked()"), self.close)
		self.connect(self.ui.saveButton, QtCore.SIGNAL("clicked()"), self.save)
		self.connect(self.ui.listsButton, QtCore.SIGNAL("clicked()"), self.get_LISTS_Folder)
		self.connect(self.ui.outputButton, QtCore.SIGNAL("clicked()"), self.get_Output_Folder)

	def save(self):
		global Oversize_Amount,LISTS_Folder,Output_Folder,Default_Tool
		Oversize_Amount = float(self.ui.oversizeBox.value())
		LISTS_Folder = str(self.ui.listsEdit.text())
		Output_Folder = str(self.ui.outputEdit.text())
		Default_Tool = int(self.ui.toolNumBox.value())
		reply = QtGui.QMessageBox.question(self, 'Save Config?', 
										 "Do you want to save these changes for next time?", QtGui.QMessageBox.Yes, QtGui.QMessageBox.No)
		if reply == QtGui.QMessageBox.Yes:
			cfg.set("Options","oversize_amount",Oversize_Amount)              
			cfg.set("Options","lists_folder",LISTS_Folder)                    
			cfg.set("Options","output_folder",Output_Folder) 
			cfg.set("Options","default_tool",Default_Tool)
			dummy=open("C:\\doorsizing.ini","w")
			cfg.write(dummy)
			dummy.close()
		self.close()

	def get_LISTS_Folder(self):
		dr = str(QtGui.QFileDialog.getExistingDirectory(self, "Select Directory"))
		if dr:
			self.ui.listsEdit.setText(dr)

	def get_Output_Folder(self):
		dr = str(QtGui.QFileDialog.getExistingDirectory(self, "Select Directory"))
		if dr:
			self.ui.outputEdit.setText(dr)

class SizingDialog(QtGui.QDialog):
	def __init__(self, parent=None):
		QtGui.QWidget.__init__(self, parent)
		self.ui = doorSizing_Ui()
		self.ui.setupUi(self)
		self.XL_fileName = ""
		self.setFixedSize(600,430)
		self.setWindowFlags(QtCore.Qt.CustomizeWindowHint | QtCore.Qt.WindowCloseButtonHint )
		self.validator = QtGui.QRegExpValidator(QtCore.QRegExp('[a-zA-Z\s]+'))

		self.delAction = QtGui.QAction("Delete Row", self)
		self.delAction.setIcon(QtGui.QIcon("remove.png"))
		self.previewAction = QtGui.QAction("Preview Door", self)

		self.popMenu = QtGui.QMenu(self)
		self.popMenu.addAction(self.delAction)

		self.ui.doorList.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
		self.ui.doorList.setSelectionBehavior(QtGui.QAbstractItemView.SelectRows)
		self.ui.doorList.setSelectionMode(QtGui.QAbstractItemView.SingleSelection)
		self.ui.doorList.setColumnWidth(1,50)
		self.ui.doorList.setColumnWidth(2,65)
		self.ui.doorList.setColumnWidth(3,65)
		self.ui.doorList.setColumnWidth(4,75)
		self.ui.doorList.setColumnWidth(5,165)

		self.connect(self.ui.doorList, QtCore.SIGNAL("customContextMenuRequested(const QPoint&)"), self.context_menu)
		self.connect(self.ui.genButton, QtCore.SIGNAL("clicked()"), self.tacos)
		self.connect(self.ui.importButton, QtCore.SIGNAL("clicked()"), self.import_xls)
		self.connect(self.ui.addButton, QtCore.SIGNAL("clicked()"), self.add_door)
		self.connect(self.ui.doorList, QtCore.SIGNAL("itemChanged (QTableWidgetItem*)"), self.data_edited)
		self.connect(self.ui.configButton, QtCore.SIGNAL("clicked()"), self.config_menu)
		self.connect(self.delAction, QtCore.SIGNAL("triggered (bool)"), self.remove_door)
		self.ui.jobNameEdit.setValidator(self.validator)
		self.ui.roomNameEdit.setValidator(self.validator)

	def config_menu(self):
		self.popup = ConfigDialog()
		self.popup.exec_()

	def data_edited(self, item):
		r = item.row()
		c = item.column()
		if c == 0:
				door[r].page = str.strip(str(item.text()))
		elif c == 1:
			try:
				door[r].pNum = int(item.text())
			except:
				QtGui.QMessageBox.warning(self, "Dude", "This has to be a whole number. No decimals or text")
		elif c == 2:
			try:
				door[r].width = float(item.text())
			except:
				QtGui.QMessageBox.warning(self, "Dude", "This has to be a number.")   
		elif c == 3:
			try:
				door[r].height = float(item.text())
			except:
				QtGui.QMessageBox.warning(self, "Dude", "This has to be a number.")  
		elif c == 4:
				door[r].lineNum = str.strip(str(item.text()))
		elif c == 5:
				door[r].jobName = str.strip(str(item.text()))

		elif c == 6:
			try:
				door[r].toolNum = int(item.text())
			except:
				QtGui.QMessageBox.warning(self, "Dude", "This has to be a number.")  
		self.refresh_list()

	def context_menu(self, point):
		self.popMenu.exec_(self.ui.doorList.mapToGlobal(point))    

	def remove_door(self):
		global doorCount
		row = int(self.ui.doorList.currentRow())
		if (row > -1) and (row <= doorCount):
			del door[row]
			doorCount -= 1
			self.refresh_list()

	def refresh_list(self):
		self.ui.doorList.blockSignals(True)
		global doorCount
		self.ui.doorList.clearContents()
		self.ui.doorList.verticalHeader().hide()
		for i in range (0, doorCount):
			self.ui.doorList.insertRow(i)
			self.ui.doorList.setItem(i,0,QtGui.QTableWidgetItem(str(door[i].page)))
			self.ui.doorList.setItem(i,1,QtGui.QTableWidgetItem(str(door[i].get_pNum())))
			self.ui.doorList.setItem(i,2,QtGui.QTableWidgetItem(str(door[i].width)))
			self.ui.doorList.setItem(i,3,QtGui.QTableWidgetItem(str(door[i].height)))
			self.ui.doorList.setItem(i,4,QtGui.QTableWidgetItem(str(door[i].lineNum)))
			self.ui.doorList.setItem(i,5,QtGui.QTableWidgetItem(str(door[i].jobName)))
			self.ui.doorList.setItem(i,6,QtGui.QTableWidgetItem(str(door[i].toolNum)))
		self.ui.doorList.setRowCount(doorCount)
		app.processEvents()
		self.ui.doorList.blockSignals(False)

	def add_door(self):
		global doorCount,Default_Tool

		name = str.strip(str(self.ui.jobNameEdit.text()))
		self.ui.jobNameEdit.setText(name)

		if str(self.ui.roomNameEdit.text()):
			room = str.strip(str(self.ui.roomNameEdit.text()))
			self.ui.roomNameEdit.setText(room)
			name = name + ' ' + room

		item = str.strip(str(self.ui.itemNumBox.text()))
		self.ui.itemNumBox.setText(item)
		door.append(doorClass(pNum=doorCount+1,
													width=float(self.ui.widthBox.value()),
													height=float(self.ui.heightBox.value()),
													lineNum=item,
													toolNum=Default_Tool,
													jobName=name))
		doorCount += 1
		self.refresh_list()

	def import_xls(self):
		global doorCount,LISTS_Folder
		self.XL_fileName = QtGui.QFileDialog.getOpenFileName(self,
																 self.tr("Open file"),
																 self.tr(LISTS_Folder),
																 self.tr("Excel Job File (*.xlsx *.xlsm)"))

		if self.XL_fileName:
			self.ui.importButton.setText("Please wait...")
			self.setWindowTitle("Please wait, crunching numbers...")
			app.processEvents()

			try:
				book = openpyxl.reader.excel.load_workbook(str(self.XL_fileName),  use_iterators = True)             
				door_sh = book.get_sheet_by_name("DOORS")
				drawer_sh = book.get_sheet_by_name("DRAWER FRONTS")
				data_sh = book.get_sheet_by_name("SHIPPING")
				pCount = 0

				for i in data_sh.iter_rows(range_string="A3"):
					#name = re.sub(r'\s+', ' ', str(i[0].internal_value))
					name = str.strip(str(i[0].internal_value))
					name = re.sub(r'([()])+', '', name)

				#Process Doors from list
				num_rows = door_sh.get_highest_row()
				start_row = 20
				for index, row in enumerate(door_sh.iter_rows()):
					if start_row < index < num_rows:
						#row 2 is width, 3 is height, 25 is Line Number
						if (row[2].internal_value) and (row[3].internal_value):
							door.append(doorClass(pNum=row[0].internal_value,
																		width=row[2].internal_value,
																		height=row[3].internal_value,
																		lineNum=str(row[25].internal_value),
																		jobName=name,
																		page=door_sh.title))
							doorCount += 1
							pCount += 1
							self.refresh_list()

				#Process Drawer Fronts from list
				num_rows = drawer_sh.get_highest_row()
				start_row = 21
				for index, row in enumerate(drawer_sh.iter_rows()):
					if start_row < index < num_rows:
						#row 2 is width, 3 is height, 11 is Line Number
						if (row[2].internal_value) and (row[3].internal_value):
							door.append(doorClass(pNum=row[0].internal_value,
																		width=row[2].internal_value,
																		height=row[3].internal_value,
																		lineNum=str(row[10].internal_value),
																		jobName=name,
																		page=drawer_sh.title))
							doorCount += 1
							pCount += 1
							self.refresh_list()

				QtGui.QMessageBox.information(self, "Success", "Processed " + str(pCount) + " doors and fronts from the Excel file.")
			except Exception as e:
				QtGui.QMessageBox.critical(self, "Error", "Unable to read data from file. Check that \n "
																									"the file is formatted correctly and not corrupted.\n"
																									"Error Message: " + repr(e))
				print "Error Text:"
				print str(self.XL_fileName)
				print repr(e)
			self.refresh_list()
			self.ui.importButton.setText("Import list from Excel file")
			self.setWindowTitle("Homestead Sizing App")
			app.processEvents()


	def tacos(self):
		global doorCount,LISTS_Folder,Output_Folder
		self.qDirName = QtGui.QFileDialog.getExistingDirectory(self,
																 self.tr("Select Job Folder"),
																 self.tr(Output_Folder),QtGui.QFileDialog.ShowDirsOnly)
		if self.qDirName:
			self.dirName = str(QtCore.QDir.fromNativeSeparators(self.qDirName) + "/")

			try:
				for i in range(0, doorCount):
					door[i].gen_file(str(self.dirName))

				QtGui.QMessageBox.information(self, "Success", "Sizing list generated successfully! \n"
																											 "File were all saved in " + self.dirName +"\n"
																											 "Then separated by the job name.")
			except Exception as e: 
				QtGui.QMessageBox.critical(self, "ERROR!!", "Something failed while generating sizing programs... \n"
																										"Error Message: " + repr(e))
				print self.dirName
				print "Error Text: " + repr(e)

if __name__ == "__main__":
	app = QtGui.QApplication(sys.argv)
	myapp = SizingDialog()
	myapp.show()
	sys.exit(app.exec_())   